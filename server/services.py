import openpyxl, os
import tensorflow as tf
import numpy as np
from models import Indicator

class ExcelParser:
    '''
    Класс для парсинга файла со всеми показателями
    '''
    def __init__(self, file_name):
        self.workbook = openpyxl.load_workbook(file_name)

    def find_rows_by_first_cell_value(self, value):
        '''
        Возвращаемый JSON:
        {
            "Data":[
                {
                    "Name": string,
                    "Years": []
                    "Indicator": string,
                    "Values": []
                },
            ]
        }
        '''
        result = {
            "Data" : []
        }

        sheet_names = self.workbook.sheetnames

        # Перебираем все (г)листы
        for sheet_index in range(1, len(sheet_names)):
            # Получаем лист по индексу
            sheet = self.workbook[sheet_names[sheet_index]]

            sheet_name = sheet.title

            # Получаем название показателя
            indicator_name = [cell.value for cell in sheet[1] if cell.value != None]

            # Получаем года
            years = [cell.value for cell in sheet[2] if cell.value != None ]

            # Приводим к единому целочисленному формату
            for i in range(len(years)):
                if isinstance(years[i], int):
                    continue
                else:
                    years[i] = int(years[i][:4])

            # Перебираем все строки в листе
            for row in sheet.iter_rows():
                # Если регион совпал
                if row[0].value == value:
                    row_data = [] # Промежуточный массив
                    interim = {
                        "Name": indicator_name[0],
                        "Years": years,
                        "Indicator": sheet_name,
                        "Values": []
                    }
                    # В цикле проверяем на число и добавляем в промежуточный массив
                    for cell in row:
                        if isinstance(cell.value, float) or isinstance(cell.value, int):
                            if isinstance(cell.value, int):
                                cell.value = float(cell.value)
                            row_data.append(cell.value)
                    # Добавляем в конечный жсон
                    interim['Values'] = row_data
                    result['Data'].append(interim)

        return result['Data']

class CURManager:
    '''
    Класс для управления данными ЦУРов
    '''
    CUR_1 = ['1', '2']
    CUR_2 = ['3', '4']
    CUR_3 = ['5', '6', '7.1', '7.2', '8', '9', '10']
    CUR_4 = ['11', '12', '13', '14', '15', '16', '17', '18.1', '18.2']
    CUR_5 = ['19', '20']
    CUR_6 = ['21', '22', '23', '24']
    CUR_7 = ['25', '26']
    CUR_8 = ['27', '28', '29', '30']
    CUR_9 = ['31', '32', '33', '34', '35', '36']
    CUR_10 = ['37', '38']
    CUR_11 = ['39', '40', '41', '42', '43', '44']
    CUR_12 = ['45']
    CUR_15 = ['47', '48']
    CUR_16 = ['50']
    CUR_17 = ['51', '52']

    DATA_CUR = {
        'CUR_1': CUR_1,
        'CUR_2': CUR_2,
        'CUR_3': CUR_3,
        'CUR_4': CUR_4,
        'CUR_5': CUR_5,
        'CUR_6': CUR_6,
        'CUR_7': CUR_7,
        'CUR_8': CUR_8,
        'CUR_9': CUR_9,
        'CUR_10': CUR_10,
        'CUR_11': CUR_11,
        'CUR_12': CUR_12,
        'CUR_13': [],
        'CUR_14':[],
        'CUR_15': CUR_15,
        'CUR_16': CUR_16,
        'CUR_17': CUR_17,
    }

    CUR_NAMES = {
        'CUR_1': "Ликвидация нищеты",
        'CUR_2': "Ликвидация голода",
        'CUR_3': "Хорошее здоровье и благополучие",
        'CUR_4': "Качественное образование",
        'CUR_5': "Гендерное равенство",
        'CUR_6': "Чистая вода и санитария",
        'CUR_7': "Недорогостоящая и чистая энергия",
        'CUR_8': "Достойная работы и экономический рост",
        'CUR_9': "Индустриализация, инновации и инфраструктура",
        'CUR_10': "Уменьшение неравенства",
        'CUR_11': "Устойчивые города и населенные пункты",
        'CUR_12': "Ответственное потребление и производство",
        'CUR_13': "Борьба с изменением климата",
        'CUR_14': "Сохранение морских экосистем",
        'CUR_15': "Сохранение экосистем суши",
        'CUR_16': "Мир, правосудие и эффективные институты",
        'CUR_17': "Партнерство в интересах устойчивого развития",
    }


class ModelExecute:

    def __init__(self, cur_id, indicator_id, values):
        self.cur_id = cur_id
        self.indicator_id = indicator_id
        self.values = values

    def process(self):

        result = self.values # здесь мы объявляем переменную со значениями, в которую потом добавим наш предикт

        indicator = Indicator.query.filter_by(id=self.indicator_id).first()

        # Подгружаем модель
        model_path = os.path.join(os.path.dirname(__file__), f'source/CUR_{self.cur_id}/indicator{indicator.number}_model')  # ну здесь понятно уже путь просто создаем
        model = tf.saved_model.load(model_path) # а тут уже передаем его

        data = np.array(self.values).reshape((-1, 1)) # дальше идет классическое преобразование данных как и при обучении

        data_tensor = tf.convert_to_tensor(data, dtype=tf.float32)
        data_tensor = tf.expand_dims(data_tensor, axis=-1)

        # Вызов метода predict модели
        output = model.signatures["serving_default"](inputs=data_tensor)

        # Получение результатов
        pred = output["output_0"].numpy()
        result.append(float(f'{pred[-1][0]:.1f}')) # вот с этим форматированием дрочь но все равно прикольно

        return result # ну и соотв. возвращаем наш массив в который мы уже дополнительно добавили предикт