'''
---------------------------------------
Парсит файл со всеми показателями
по фильтру на Ростовскую область
---------------------------------------
'''

import openpyxl

def find_rows_by_first_cell_value(file_name, value):
    '''
    Возвращаемый JSON:
    {
        "Data":[
            {
                "Name": string,
                "Years": []
                "Indicator": string,
                "Values": []
            },
        ]
    }
    '''
    # Загружаем файл
    workbook = openpyxl.load_workbook(file_name)

    # Создание result для конечных данных
    result = {
        "Data" : []
    }

    sheet_names = workbook.sheetnames

    # Перебираем все (г)листы
    for sheet_index in range(1, len(sheet_names)):
        # Получаем лист по индексу
        sheet = workbook[sheet_names[sheet_index]]

        sheet_name = sheet.title

        # Получаем название показателя
        indicator_name = [cell.value for cell in sheet[1] if cell.value != None]

        # Получаем года
        years = [cell.value for cell in sheet[2] if cell.value != None ]

        # Приводим к единому целочисленному формату
        for i in range(len(years)):
            if isinstance(years[i], int):
                continue
            else:
                years[i] = int(years[i][:4])

        # Перебираем все строки в листе
        for row in sheet.iter_rows():
            # Если регион совпал
            if row[0].value == value:
                row_data = [] # Промежуточный массив
                interim = {
                    "Name": indicator_name[0],
                    "Years": years,
                    "Indicator": sheet_name,
                    "Values": []
                }
                # В цикле проверяем на число и добавляем в промежуточный массив
                for cell in row:
                    if isinstance(cell.value, float) or isinstance(cell.value, int):
                        if isinstance(cell.value, int):
                            cell.value = float(cell.value)
                        row_data.append(cell.value)
                # Добавляем в конечный жсон
                interim['Values'] = row_data
                result['Data'].append(interim)

    return result['Data']
